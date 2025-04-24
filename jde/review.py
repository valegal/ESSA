
import time
import os
import shutil
import re
import pdfplumber
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from config import config

# Expresiones regulares mejoradas
regex_agrupacion = re.compile(r'EMONTANC.*?\s(\d{5})')
regex_carga = re.compile(r'(\d{5})\s+\d{4}/\d{2}/\d{2}')
regex_fecha_contable = re.compile(r'(\d{4}/\d{2}/\d{2})\s+.*Asientos Interface Facturacion')
regex_debitos = re.compile(r'DEBITOS GENERAL\s+([\d,]+\.\d{2})')
regex_creditos = re.compile(r'CREDITOS GENERAL\s+([\d,]+\.\d{2})-?')

regex_nfase = re.compile(r'(\d{4}/\d{2}/\d{2})\s+(\d{4}/\d{2}/\d{2})\s+(\d)\b')

# Carpeta con los archivos PDF
carpeta_pdf = config.FOLDER_R5609FCT
carpeta_pdf2 = config.FOLDER_DINAMICAS
carpeta_origen = str(config.FOLDER_ORIGEN)

#-----------------------------------------------------------------------------

def review_pdfs(driver, reportes, batchcarga):
    """
    Esta función automatiza la revisión de los Reportes para Generar Dinámica Contable, seleccionando los archivos para que se genere el pdf.
    """

    # Hacer clic en el icono de la tabla
    icono = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='listRRpt_WSJ']/table/tbody/tr/td[1]"))
    )
    icono.click()
    time.sleep(5)

    # Cambiar al iframe
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "e1menuAppIframe")))
    driver.switch_to.frame(driver.find_element(By.ID, "e1menuAppIframe"))

    for i in range(reportes - 1, -1, -1):  
        # Esperar a que el elemento esté presente y visible
        tarea_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, f"//*[@id='G0_1_R{i}']/td[1]/div/input"))
        )

        # Crear una instancia de ActionChains y hacer doble clic en el elemento
        action = ActionChains(driver)
        action.double_click(tarea_element).perform()

        print(f"Doble clic realizado en el elemento {i}")
        time.sleep(3)
    
    time.sleep(3)

    # Mover los archivos descargados
    mover_reportes()
    valores_columna_dos = contrastar_debitos_y_creditos(carpeta_pdf, batchcarga)
    time.sleep(5)
    eliminar_reportes()
    return valores_columna_dos


#-----------------------------------------------------------------------------

def mover_reportes():
    """
    Esta función automatiza el proceso de mover todos los archivos PDF cuyo nombre inicie por 'R5609FCT_'
    desde la carpeta Descargas a las carpetas R5609FCT y ReportesDinamicaContable.
    """
    # Definir rutas
    carpeta_origen = config.FOLDER_ORIGEN
    carpeta_destino = carpeta_origen / "R5609FCT"
    carpeta_destino2 = carpeta_origen / "ReportesDinamicaContable"

    # Asegurar que las carpetas de destino existen
    for carpeta in [carpeta_destino, carpeta_destino2]:
        carpeta.mkdir(parents=True, exist_ok=True)

    # Buscar archivos que cumplan con el patrón
    for archivo in carpeta_origen.glob("R5609FCT_*.pdf"):
        # Mover el archivo a la carpeta R5609FCT
        destino_1 = carpeta_destino / archivo.name
        shutil.move(str(archivo), str(destino_1))
        print(f"Movido a reportes: {archivo.name}")

        # Copiar a la carpeta ReportesDinamicaContable
        destino_2 = carpeta_destino2 / archivo.name
        shutil.copy(str(destino_1), str(destino_2))

    print("✅ Proceso completado.")


#-----------------------------------------------------------------------------

def contrastar_debitos_y_creditos(carpeta_pdf, batchcarga):
    """
    Revisa todos los PDFs en la carpeta y extrae información clave.
    """
        
    resultados = []

    for archivo in os.listdir(carpeta_pdf):
        if archivo.endswith(".pdf"):
            pdf_path = os.path.join(carpeta_pdf, archivo)
            with pdfplumber.open(pdf_path) as pdf:
                primera_pagina = pdf.pages[0].extract_text()
                ultima_pagina = pdf.pages[-1].extract_text()

                if not primera_pagina or not ultima_pagina:
                    print(f"⚠ No se pudo extraer texto de {archivo}")
                    continue

                # Extraer información clave
                agrupacion = regex_agrupacion.search(primera_pagina)
                fecha_contable = regex_fecha_contable.search(primera_pagina)
                match_debitos = regex_debitos.search(ultima_pagina)
                match_creditos = regex_creditos.search(ultima_pagina)
                fase_carga = regex_nfase.search(primera_pagina)

                if not match_debitos or not match_creditos or not agrupacion or not fecha_contable or not fase_carga:
                    print(f"⚠ Datos faltantes en {archivo}")
                    continue

                # Convertir los valores correctamente
                debitos = float(match_debitos.group(1).replace(",", ""))
                creditos = float(match_creditos.group(1).replace(",", ""))

                # Verificar igualdad
                if debitos == abs(creditos):
                    print(f"✅ Débitos y créditos coinciden en {archivo}")
                    resultados.append({
                            "archivo": archivo,   
                            "agrupacion": agrupacion.group(1),
                            "fecha_contable": fecha_contable.group(1) if fecha_contable else "N/A",
                            "debitos": debitos,
                            "creditos": creditos,
                            "fecha_contable": fecha_contable.group(1),
                            "fase_carga": fase_carga.group(3)
                    })
                else:
                    print(f"⚠ Descuadre en {archivo}: Débitos {debitos} vs Créditos {creditos}")

    # Crear el diccionario con la correspondencia correcta
    columna_dos = {r["fase_carga"]: r["agrupacion"] for r in resultados}
    
    # Actualizar el Excel con los valores correctos
    update_excel_batch_agrupacion(resultados, batchcarga)
    return columna_dos


#-----------------------------------------------------------------------------
def update_excel_batch_agrupacion(resultados, batchcarga):
    excel_path = config.EXCEL_PATH
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        print("✅ Archivo Excel cargado correctamente.")
    except FileNotFoundError:
        print(f"❌ Error: El archivo {excel_path} no se encontró.")
        return {}
    
    hojas_map = {
        "1": "1-FACTURACIÓN",
        "2": "2-AUTOCONSUMOS",
        "3": "3-AJUSTES",
        "4": "4-RECAUDOS",
        "5": "5-CASTIGO"
    }

    columna = {}
    
    for resultado in resultados:
        fase_carga = resultado["fase_carga"]
        agrupacion = resultado["agrupacion"]
        fecha = resultado["fecha_contable"]

        if fase_carga not in batchcarga:
            print(f"⚠ Fase de carga {fase_carga} no encontrada en batchcarga")
            continue

        hoja_nombre = hojas_map.get(fase_carga)
        if not hoja_nombre or hoja_nombre not in wb.sheetnames:
            print(f"⚠ No se encontró la hoja {hoja_nombre} en el archivo Excel.")
            continue

        try:
            dia = int(fecha.split("/")[2])
            celda_destino = f"C{7 + dia}"
        except ValueError:
            print(f"⚠ Fecha inválida en {resultado['archivo']}: {fecha}")
            continue

        hoja = wb[hoja_nombre]
        hoja[celda_destino] = agrupacion
        print(f"✅ Actualizado {hoja_nombre} en {celda_destino} con {agrupacion}")
    
    try:
        wb = openpyxl.load_workbook(str(excel_path))
        wb.close()
        print("✅ Archivo Excel actualizado correctamente.")
    except Exception as e:
        print(f"❌ Error al guardar el archivo Excel: {e}")
    
    return columna




#-----------------------------------------------------------------------------

from pathlib import Path

def eliminar_reportes():
    """
    Elimina los archivos PDF de la carpeta 'R5609FCT' después del proceso.
    """
    carpeta_destino = config.FOLDER_R5609FCT 

    if carpeta_destino.exists():
        for archivo in carpeta_destino.iterdir():
            if archivo.is_file():
                archivo.unlink()
                print(f"🗑 Eliminado: {archivo.name}")
        print("✅ Todos los archivos han sido eliminados.")
    else:
        print(f"⚠ La carpeta {carpeta_destino} no existe.")
        
        
        
        
#======================= REVIEW PDF ÚNICO ====================



def review_pdf_unico(driver, reportes, batchcarga):
    """
    Esta función automatiza la revisión de los Reportes para Generar Dinámica Contable, seleccionando los archivos para que se genere el pdf.
    """

    # Hacer clic en el icono de la tabla
    icono = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='listRRpt_WSJ']/table/tbody/tr/td[1]"))
    )
    icono.click()
    time.sleep(5)

    # Cambiar al iframe
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "e1menuAppIframe")))
    driver.switch_to.frame(driver.find_element(By.ID, "e1menuAppIframe"))

    for i in range(reportes - 1, -1, -1):  
        # Esperar a que el elemento esté presente y visible
        tarea_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, f"//*[@id='G0_1_R{i}']/td[1]/div/input"))
        )

        # Crear una instancia de ActionChains y hacer doble clic en el elemento
        action = ActionChains(driver)
        action.double_click(tarea_element).perform()

        print(f"Doble clic realizado en el elemento {i}")
        time.sleep(3)
    
    time.sleep(3)

    # Mover los archivos descargados
    mover_reporte()
    valores_columna_dos = contrastar_debito_y_credito(carpeta_pdf, batchcarga)
    time.sleep(5)
    eliminar_reporte()
    return valores_columna_dos


#-----------------------------------------------------------------------------

def mover_reporte():
    """
    Esta función automatiza el proceso de mover todos los archivos PDF cuyo nombre inicie por 'R5609FCT_'
    desde la carpeta Descargas a las carpetas R5609FCT y ReportesDinamicaContable.
    """
    # Definir rutas
    carpeta_origen = config.FOLDER_ORIGEN
    carpeta_destino = carpeta_origen / "R5609FCT"
    carpeta_destino2 = carpeta_origen / "ReportesDinamicaContable"

    # Asegurar que las carpetas de destino existen
    for carpeta in [carpeta_destino, carpeta_destino2]:
        carpeta.mkdir(parents=True, exist_ok=True)

    # Buscar archivos que cumplan con el patrón
    for archivo in carpeta_origen.glob("R5609FCT_*.pdf"):
        # Mover el archivo a la carpeta R5609FCT
        destino_1 = carpeta_destino / archivo.name
        shutil.move(str(archivo), str(destino_1))
        print(f"Movido a reportes: {archivo.name}")

        # Copiar a la carpeta ReportesDinamicaContable
        destino_2 = carpeta_destino2 / archivo.name
        shutil.copy(str(destino_1), str(destino_2))

    print("✅ Proceso completado.")


#-----------------------------------------------------------------------------

def contrastar_debito_y_credito(carpeta_pdf, batchcarga):
    """
    Revisa todos los PDFs en la carpeta y extrae información clave.
    """
        
    resultados = []

    for archivo in os.listdir(carpeta_pdf):
        if archivo.endswith(".pdf"):
            pdf_path = os.path.join(carpeta_pdf, archivo)
            with pdfplumber.open(pdf_path) as pdf:
                primera_pagina = pdf.pages[0].extract_text()
                ultima_pagina = pdf.pages[-1].extract_text()

                if not primera_pagina or not ultima_pagina:
                    print(f"⚠ No se pudo extraer texto de {archivo}")
                    continue

                # Extraer información clave
                agrupacion = regex_agrupacion.search(primera_pagina)
                fecha_contable = regex_fecha_contable.search(primera_pagina)
                match_debitos = regex_debitos.search(ultima_pagina)
                match_creditos = regex_creditos.search(ultima_pagina)
                fase_carga = regex_nfase.search(primera_pagina)

                if not match_debitos or not match_creditos or not agrupacion or not fecha_contable or not fase_carga:
                    print(f"⚠ Datos faltantes en {archivo}")
                    continue

                # Convertir los valores correctamente
                debitos = float(match_debitos.group(1).replace(",", ""))
                creditos = float(match_creditos.group(1).replace(",", ""))

                # Verificar igualdad
                if debitos == abs(creditos):
                    print(f"✅ Débitos y créditos coinciden en {archivo}")
                    resultados.append({
                            "archivo": archivo,   
                            "agrupacion": agrupacion.group(1),
                            "fecha_contable": fecha_contable.group(1) if fecha_contable else "N/A",
                            "debitos": debitos,
                            "creditos": creditos,
                            "fecha_contable": fecha_contable.group(1),
                            "fase_carga": fase_carga.group(3)
                    })
                else:
                    print(f"⚠ Descuadre en {archivo}: Débitos {debitos} vs Créditos {creditos}")

    # Crear el diccionario con la correspondencia correcta
    columna_dos = {r["fase_carga"]: r["agrupacion"] for r in resultados}
    
    # Actualizar el Excel con los valores correctos
    update_excel_batch_agrupacion_uno(resultados, batchcarga)
    return columna_dos


#-----------------------------------------------------------------------------
def update_excel_batch_agrupacion_uno(resultados, batchcarga):
    excel_path = config.EXCEL_PATH
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        print("✅ Archivo Excel cargado correctamente.")
    except FileNotFoundError:
        print(f"❌ Error: El archivo {excel_path} no se encontró.")
        return {}
    
    hojas_map = {
        "1": "1-FACTURACIÓN",
        "2": "2-AUTOCONSUMOS",
        "3": "3-AJUSTES",
        "4": "4-RECAUDOS",
        "5": "5-CASTIGO"
    }

    columna = {}
    
    for resultado in resultados:
        fase_carga = resultado["fase_carga"]
        agrupacion = resultado["agrupacion"]
        fecha = resultado["fecha_contable"]

        if fase_carga not in batchcarga:
            print(f"⚠ Fase de carga {fase_carga} no encontrada en batchcarga")
            continue

        hoja_nombre = hojas_map.get(fase_carga)
        if not hoja_nombre or hoja_nombre not in wb.sheetnames:
            print(f"⚠ No se encontró la hoja {hoja_nombre} en el archivo Excel.")
            continue

        try:
            dia = int(fecha.split("/")[2])
            celda_destino = f"C{7 + dia}"
        except ValueError:
            print(f"⚠ Fecha inválida en {resultado['archivo']}: {fecha}")
            continue

        hoja = wb[hoja_nombre]
        hoja[celda_destino] = agrupacion
        print(f"✅ Actualizado {hoja_nombre} en {celda_destino} con {agrupacion}")
    
    try:
        wb = openpyxl.load_workbook(str(excel_path))
        wb.close()
        print("✅ Archivo Excel actualizado correctamente.")
    except Exception as e:
        print(f"❌ Error al guardar el archivo Excel: {e}")
    
    return columna




#-----------------------------------------------------------------------------

from pathlib import Path

def eliminar_reporte():
    """
    Elimina los archivos PDF de la carpeta 'R5609FCT' después del proceso.
    """
    carpeta_destino = config.FOLDER_R5609FCT  # Ya es un Path

    if carpeta_destino.exists():
        for archivo in carpeta_destino.iterdir():
            if archivo.is_file():
                archivo.unlink()
                print(f"🗑 Eliminado: {archivo.name}")
        print("✅ Todos los archivos han sido eliminados.")
    else:
        print(f"⚠ La carpeta {carpeta_destino} no existe.")
        